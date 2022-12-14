import openpyxl
from openpyxl.styles import PatternFill

def rowcount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def columncount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def readData(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum,columnnum).value

def writeData(file, sheetname, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,columnnum).value = data
    workbook.save(file)

def green(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    sheet.cell(rownum, columnnum).fill = greenfill
    workbook.save(file)

def red(file, sheetname, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rownum, columnnum).fill = redfill
    workbook.save(file)






